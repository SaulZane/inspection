from openpyxl import load_workbook
import pandas as pd
from colorama import Fore, Back, Style, init

init(autoreset=True)

try:
    print(
        Fore.GREEN
        + "-----欢迎使用社会机构查验非本行政区嫌疑程序导出软件-----©张硕 保留所有权利|2024.3.1|V1.0-----"
    )
    print(
        Fore.CYAN
        + "使用说明：输入表名称必须是“表2.xlsx”，必须有“基础表”数据页。并且必须在程序的同名文件夹下"
    )
    print(Fore.CYAN + "必须含有“社会机构所在区 社会机构名称 登记地址所属区” 三个字段，表头名字不能变")
    print(Fore.CYAN + "如果运行成功，在同名文件夹下会看到result.xlsx结果表")
    print(Fore.GREEN + "按任意键开始运行程序...")
    input()
    print(Fore.RED + "--读表2.xlsx中数据，请稍等，不要动电脑--")

    # 使用pandas库读取excel文件
    data = pd.read_excel("表2.xlsx", sheet_name="基础表", header=0)
    # data= data.head(500)
    # 将数据转换为字典
    df = pd.DataFrame.from_records(data)
    # 将数据转换为字典
    records = df.to_dict(orient="records")

    def input_with_default_threshold(prompt):
        user_input = input(f"{prompt}  ")
        if user_input.strip() == "":
            return 0.5
        return user_input

    threshold = 0.5
    print(Fore.RED + "--读取完成，请输入内容--")
    threshold = float(
        input_with_default_threshold(Fore.GREEN + "请输入占比(默认0.5)：")
    )
    print(Fore.RED + "--跑数据中，以下为调试信息--")

    def find_violations(records, threshold=0.5):

        violations = []
        data = []
        for record in records:

            i = record["社会机构所在区"]
            u = record["社会机构名称"]
            b = record["登记地址所属区"]
            key = (u, i, b)
            data.append(key)

        # 1、以上计算相同内容出现次数
        # 定义 result 列表
        result = []

        # 定义一个字典，用来存储每个元组在 data 列表中出现的次数
        count = {}

        # 遍历 data 列表中的每个元组
        for tup in data:
            # 判断 tup 是否在 count 字典中
            if tup in count:
                # 如果在，就把它的次数加一
                count[tup] += 1
            else:
                # 如果不在，就把它的次数设为一
                count[tup] = 1

        # 遍历 count 字典中的每个键值对
        for key, value in count.items():
            # 把 key 和 value 组合成一个新的元组，添加到 result 列表中
            result.append(key + (value,))

        # 2、合并同类项
        data = result
        result = []

        sum_dict = {}

        # 遍历data列表
        for item in data:
            key = (item[0], item[1])  # 使用元素中的前两个值作为键
            value = {item[2]: item[3]}  # 使用元素中的第三个和第四个值构建字典
            if key in sum_dict:
                sum_dict[key].update(value)  # 将字典更新到已有键的值中
            else:
                sum_dict[key] = value  # 初始化字典

        # 创建结果列表
        result = []

        # 遍历字典中的键值对，构建结果
        for key, values in sum_dict.items():

            sums = sum(values.values())  # 计算总数
            subresult = (key[0], key[1], values, sums)
            result.append(subresult)

        # print(result)
        # 3、找到违规项
        violations = []
        for item in result:
            tempkeyvalue = dict(item[2])
            for key, value in tempkeyvalue.items():
                if value >= item[3] * threshold and key != item[1]:
                    if (item) not in violations:
                        violations.append(item)
        print(violations)

        return violations

    # 输出结果，新建一个excel文件，将结果写入到excel文件中，名称为result.xlsx
    result = find_violations(records, threshold)

    # 创建DataFrame
    df = pd.DataFrame(
        result,
        columns=["社会机构名称", "社会机构所在区", "查验登记地址所属区", "总查验量"],
    )

    # 创建ExcelWriter对象
    writer = pd.ExcelWriter("result.xlsx", engine="openpyxl")

    # 将数据写入表1
    df.to_excel(writer, sheet_name="基本表", index=False)

    # 展开字典数据为新的列
    df_expanded = pd.concat(
        [
            df.drop(["查验登记地址所属区"], axis=1),
            df["查验登记地址所属区"].apply(pd.Series),
        ],
        axis=1,
    )

    # 将数据写入表2
    df_expanded.to_excel(writer, sheet_name="展开表", index=False)

    # 保存Excel文件
    writer.close()
    # 写标记值
    # 读取Excel文件

    # 读取Excel文件
    file_path = "result.xlsx"
    workbook = load_workbook(file_path)

    # 选择要写入数据的工作表
    sheet_name = "基本表"  # 假设要写入的工作表名称为'Sheet1'
    sheet = workbook[sheet_name]

    # 获取要写入数据的行和列索引
    row_index = 2
    col_index = 7

    # 写入数据到指定单元格
    cell = sheet.cell(row=row_index, column=col_index)
    cell.value = "超出总量的" + str(threshold * 100) + "%为违规项"

    # 保存修改后的Excel文件
    workbook.save(file_path)
    print("-------")
    print(
        Fore.RED + "恭喜！！result.xlsx文件已生成，可以操作电脑了!按任意键退出程序..."
    )
    input()

except Exception as e:
    print(
        Fore.RED
        + "程序出现异常，请检查表格或者输入是否符合要求，或者联系开发者"
        + str(e.args)
    )
    input()
